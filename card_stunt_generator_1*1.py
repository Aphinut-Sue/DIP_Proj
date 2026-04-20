import cv2
import numpy as np
import matplotlib.pyplot as plt
import os
import json
import string

from openpyxl import Workbook
from openpyxl.styles import Alignment

# import metrics.py functions
from metrics import evaluate, save_report

# -----------------------------
# CONFIG
# -----------------------------
TOTAL_COLORS = 50

PLATE_ROWS = 25
PLATE_COLS = 50

ROW_LABELS = list(string.ascii_uppercase[:PLATE_ROWS])


# -----------------------------
# LOAD IMAGE
# -----------------------------
def load_image(path):

    img = cv2.imread(path)

    if img is None:
        raise FileNotFoundError(path)

    img = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)

    return img


# -----------------------------
# ENHANCE CONTRAST (CLAHE)
# -----------------------------
def enhance_contrast(img):

    lab = cv2.cvtColor(img, cv2.COLOR_RGB2LAB)

    l, a, b = cv2.split(lab)

    clahe = cv2.createCLAHE(
        clipLimit=3.0,
        tileGridSize=(8,8)
    )

    l = clahe.apply(l)

    lab = cv2.merge((l, a, b))

    return cv2.cvtColor(lab, cv2.COLOR_LAB2RGB)


# -----------------------------
# EDGE PRESERVING SMOOTH
# -----------------------------
def edge_preserve(img):

    return cv2.edgePreservingFilter(
        img,
        flags=1,
        sigma_s=60,
        sigma_r=0.4
    )


# -----------------------------
# SOFT SHARPEN
# -----------------------------
def sharpen_soft(img):

    smooth = cv2.GaussianBlur(img, (0,0), 1.5)

    return cv2.addWeighted(
        img, 1.4,
        smooth, -0.4,
        0
    )


# -----------------------------
# DENOISE
# -----------------------------
def denoise_image(img):

    return cv2.bilateralFilter(img, 9, 75, 75)


# -----------------------------
# LOAD COLOR BOOK
# -----------------------------
def load_color_book_json(path):

    with open(path) as f:
        data = json.load(f)

    palette_rgb = []
    color_book = []

    for i, c in enumerate(data["colors"]):

        hex_color = c["hex"].lstrip("#")

        r = int(hex_color[0:2], 16)
        g = int(hex_color[2:4], 16)
        b = int(hex_color[4:6], 16)

        palette_rgb.append([r, g, b])

        color_book.append({
            "color_id": c["color_id"],
            "rgb": (r, g, b),
            "hex": c["hex"],
            "index": i
        })

    palette_rgb = np.array(palette_rgb, dtype=np.uint8)

    palette_lab = cv2.cvtColor(
        palette_rgb.reshape(1, -1, 3),
        cv2.COLOR_RGB2LAB
    ).reshape(-1, 3)

    return palette_rgb, palette_lab, color_book


# -----------------------------
# MAP IMAGE TO PALETTE
# -----------------------------
def map_to_palette(img, palette_rgb, palette_lab, batch_size=5000):

    img_lab = cv2.cvtColor(img, cv2.COLOR_RGB2LAB)

    h, w, _ = img_lab.shape

    pixels = img_lab.reshape(-1, 3).astype(np.int16)

    indices_list = []

    for i in range(0, len(pixels), batch_size):

        chunk = pixels[i:i+batch_size]

        dists = np.linalg.norm(
            chunk[:, None, :] - palette_lab[None, :, :],
            axis=2
        )

        indices = np.argmin(dists, axis=1)

        indices_list.append(indices)

    color_indices = np.concatenate(indices_list)
    
    quantized = palette_rgb[color_indices].reshape((h, w, 3))

    return quantized, color_indices.reshape((h, w))


# -----------------------------
# COLOR STATISTICS
# -----------------------------
def get_top_colors(color_indices, color_book, top_n):

    flat = color_indices.flatten()

    counts = np.bincount(flat, minlength=len(color_book))
    total_pixels = flat.size

    stats = []

    for i, count in enumerate(counts):

        stats.append({
            "color_id": color_book[i]["color_id"],
            "hex": color_book[i]["hex"],
            "rgb": color_book[i]["rgb"],
            "pixels": int(count),
            "ratio_percent": count / total_pixels * 100,
            "orig_index": i
        })

    stats.sort(key=lambda x: x["pixels"], reverse=True)

    return stats[:top_n], stats


# -----------------------------
# COLOR STATS FROM OUTPUT IMAGE
# -----------------------------
def get_color_stats_from_image(img, color_book):

    pixels = img.reshape(-1,3)

    unique, counts = np.unique(pixels, axis=0, return_counts=True)

    total_pixels = pixels.shape[0]

    stats = []

    for rgb, count in zip(unique, counts):

        for c in color_book:
            if tuple(rgb) == c["rgb"]:

                stats.append({
                    "color_id": c["color_id"],
                    "hex": c["hex"],
                    "rgb": c["rgb"],
                    "pixels": int(count),
                    "ratio_percent": count / total_pixels * 100
                })

                break

    stats.sort(key=lambda x: x["pixels"], reverse=True)

    return stats


# -----------------------------
# PRINT COLOR TABLE
# -----------------------------
def print_color_table(color_stats):

    print("\nColor Usage Statistics\n")

    print("+----+---------+--------+--------+---------+")

    print("|Rank|Color ID | HEX    | Pixels | Percent |")

    print("+----+---------+--------+--------+---------+")

    rank = 1

    for c in color_stats:

        if c["pixels"] == 0:
            continue

        print(
            f"|{rank:>4}"
            f"|{c['color_id']:>9}"
            f"|{c['hex']:>8}"
            f"|{c['pixels']:>8}"
            f"|{c['ratio_percent']:>7.3f}% |"
        )

        rank += 1

    print("+----+---------+--------+--------+---------+")


# -----------------------------
# CREATE CARD GRID
# -----------------------------
def create_card_grid(color_indices, palette):

    return palette[color_indices]


# -----------------------------
# SPLIT INTO PLATES
# -----------------------------
def split_into_plates(color_indices, cards_per_row, cards_per_column):

    plates = color_indices.reshape(
        PLATE_ROWS,
        cards_per_row,
        PLATE_COLS,
        cards_per_column
    ).swapaxes(1,2)

    return plates


# -----------------------------
# DISPLAY GRID
# -----------------------------
def display_card_grid(card_grid, show_plate_lines=True):

    plt.figure(figsize=(TOTAL_COLS/20, TOTAL_ROWS/20))

    plt.imshow(card_grid.astype(np.uint8))

    if show_plate_lines:

        for y in range(0, TOTAL_ROWS, cards_per_row):
            plt.axhline(y - 0.5, linewidth=0.3, color='black', alpha=0.5)

        for x in range(0, TOTAL_COLS, cards_per_column):
            plt.axvline(x - 0.5, linewidth=0.3, color='black', alpha=0.5)

    xticks = np.arange(cards_per_column/2, TOTAL_COLS, cards_per_column)
    yticks = np.arange(cards_per_row/2, TOTAL_ROWS, cards_per_row)

    plate_title = f"1:{cards_per_row * cards_per_column}"

    plt.xticks(xticks, range(1, PLATE_COLS + 1))
    plt.yticks(yticks, ROW_LABELS)

    plt.xlabel("Column (1–50)")
    plt.ylabel("Row (A–Y)")
    plt.title(f"Card Stunt Grid ({plate_title})")

    plt.show()


# -----------------------------
# DISPLAY SPECIFIC PLATE
# -----------------------------
def display_specific_plate(plates, color_book, position):

    position = position.upper().strip()

    row_letter = ''.join([c for c in position if c.isalpha()])
    col_num = ''.join([c for c in position if c.isdigit()])

    if not row_letter or not col_num:
        print("Invalid plate")
        return

    try:

        row_idx = ROW_LABELS.index(row_letter)
        col_idx = int(col_num) - 1

    except:
        print("Invalid plate")
        return

    plate = plates[row_idx][col_idx]

    fig, ax = plt.subplots(figsize=(6,6))

    plate_img = np.zeros((cards_per_row, cards_per_column, 3), dtype=np.uint8)

    for r in range(cards_per_row):
        for c in range(cards_per_column):

            idx = plate[r,c]

            rgb = color_book[idx]["rgb"]

            plate_img[r,c] = rgb

            color = "white" if np.mean(rgb) < 128 else "black"

            ax.text(
                c, r, str(idx+1), # +1 for 1-50
                ha="center", va="center",
                color=color,
                fontsize=14,
                fontweight="bold"
            )

    ax.imshow(plate_img)
    ax.set_title(f"Plate {position}")
    ax.axis("off")

    plt.show()


# -----------------------------
# EXPORT PLATES TO EXCEL
# -----------------------------
def export_plates_to_excel(plates, cards_per_row, cards_per_column, filename="plates.xlsx"):

    wb = Workbook()
    ws = wb.active
    ws.title = "Plates"

    row_cursor = 1

    for r in range(PLATE_ROWS):

        # ----- HEADER -----
        col_cursor = 1

        for c in range(PLATE_COLS):

            label = f"-{ROW_LABELS[r]}{c+1}-"

            ws.cell(row=row_cursor, column=col_cursor, value=label).alignment = Alignment(horizontal="center", vertical="center")

            col_cursor += cards_per_column + 1

        row_cursor += 1


        # ----- ROWS OF PLATE -----
        for pr in range(cards_per_row):

            col_cursor = 1

            for c in range(PLATE_COLS):

                plate = plates[r][c]

                for pc in range(cards_per_column):

                    ws.cell(
                        row=row_cursor,
                        column=col_cursor + pc,
                        value=int(plate[pr][pc]+1)
                    )

                col_cursor += cards_per_column + 1

            row_cursor += 1

        row_cursor += 1  # space between rows

    wb.save(filename)

    print(f"\nExcel exported → {filename}")


# -----------------------------
# USER INPUT
# -----------------------------
def get_user_image(base_dir):

    IMAGE_EXTENSIONS = (
        ".png",".jpg",".jpeg",".bmp",".tiff",".tif",
        ".webp",".gif",".jp2",".ppm",".pgm",".pbm"
    )

    files = [
        f for f in os.listdir(base_dir)
        if f.lower().endswith(IMAGE_EXTENSIONS)
    ]

    if not files:
        print("No image files found.")
        return None

    files.sort()

    print("\nAvailable images\n")

    for i, f in enumerate(files, start=1):
        print(f"{i}) {f}")

    while True:

        try:
            choice = int(input("\nSelect image (1-{}): ".format(len(files))))

            if 1 <= choice <= len(files):
                return os.path.join(base_dir, files[choice - 1])

        except:
            pass

        print("Invalid selection")


# -----------------------------
# CARD LAYOUT OPTION
# -----------------------------
def get_card_layout():

    print("\nSelect card layout\n")
    print("1) 1:16  (4 x 4)")
    print("2) 1:20  (4 x 5)")
    print("3) 1:25  (5 x 5)")

    while True:

        try:
            choice = int(input("\nSelect option (1-3): "))

            if choice == 1:
                return 4,4

            elif choice == 2:
                return 4,5

            elif choice == 3:
                return 5,5

        except:
            pass

        print("Invalid selection")


# -----------------------------
# GET USER TOP N COLORS
# -----------------------------
def get_user_top_n(max_colors):

    while True:

        try:
            n = int(input(f"Enter number of colors (1-{max_colors}): "))
            if 1 <= n <= max_colors:
                return n
        except:
            pass

        print("Invalid input")


# -----------------------------
# SAVE IMAGE (PIXEL SCALE)
# -----------------------------
def save_image(img, path, scale=10):

    os.makedirs(os.path.dirname(path), exist_ok=True)

    h, w = img.shape[:2]

    enlarged = cv2.resize(
        img,
        (w*scale, h*scale),
        interpolation=cv2.INTER_NEAREST
    )

    img_bgr = cv2.cvtColor(enlarged.astype(np.uint8), cv2.COLOR_RGB2BGR)

    cv2.imwrite(path, img_bgr)

    print(f"Image saved → {path}")


# -----------------------------
# MAIN
# -----------------------------
if __name__ == "__main__":

    # =====================================================
    # 1. PATH SETUP
    # =====================================================
    BASE_DIR = os.path.dirname(os.path.abspath(__file__))

    COLOR_BOOK_PATH = os.path.join(BASE_DIR, "color_book.json")
    OUTPUT_DIR = os.path.join(BASE_DIR, "output")
    PERFORMANCE_DIR = os.path.join(BASE_DIR, "performance")
    IMAGE_DIR = os.path.join(BASE_DIR, "images")

    if not os.path.exists(COLOR_BOOK_PATH):
        raise FileNotFoundError(f"Color book not found: {COLOR_BOOK_PATH}")


    # =====================================================
    # 2. USER INPUT
    # =====================================================
    IMAGE_PATH = get_user_image(IMAGE_DIR)

    # =====================================================
    # CARD LAYOUT
    # =====================================================
    cards_per_row, cards_per_column = get_card_layout()

    TOTAL_ROWS = PLATE_ROWS * cards_per_row
    TOTAL_COLS = PLATE_COLS * cards_per_column

    # =====================================================
    # TOP_N
    # =====================================================
    top_n = get_user_top_n(TOTAL_COLORS)

    # =====================================================
    # 3. IMAGE PREPROCESSING
    # =====================================================
    img = load_image(IMAGE_PATH)

    # contrast enhancement
    img = enhance_contrast(img)

    # light denoise
    img = cv2.bilateralFilter(img, 7, 50, 50)

    # sharpen
    smooth = cv2.GaussianBlur(img, (0,0), 1.0)
    img = cv2.addWeighted(img, 1.3, smooth, -0.3, 0)

    # slight blur before quantization
    img = cv2.GaussianBlur(img, (3,3), 0.4)


    # =====================================================
    # 4. RESIZE IMAGE TO GRID SIZE
    # =====================================================
    img_small = cv2.resize(
        img,
        (TOTAL_COLS, TOTAL_ROWS),
        interpolation=cv2.INTER_AREA
    )


    # =====================================================
    # 5. LOAD COLOR BOOK
    # =====================================================
    palette_rgb, palette_lab, color_book = load_color_book_json(COLOR_BOOK_PATH)


    # =====================================================
    # 6. INITIAL QUANTIZATION (ALL COLORS)
    # =====================================================
    reduced_50, indices_50 = map_to_palette(
        img_small,
        palette_rgb,
        palette_lab
    )


    # =====================================================
    # 7. SELECT TOP COLORS
    # =====================================================
    top_colors, _ = get_top_colors(indices_50, color_book, top_n)

    top_palette = np.array(
        [c["rgb"] for c in top_colors],
        dtype=np.uint8
    )

    top_palette_lab = cv2.cvtColor(
        top_palette.reshape(1,-1,3),
        cv2.COLOR_RGB2LAB
    ).reshape(-1,3)

    index_map = np.array(
        [c["orig_index"] for c in top_colors]
    )


    # =====================================================
    # 8. FINAL QUANTIZATION
    # =====================================================
    reduced_img, color_indices = map_to_palette(
        img_small,
        top_palette,
        top_palette_lab
    )

    mapped_indices = index_map[color_indices]


    # =====================================================
    # 9. COLOR STATISTICS
    # =====================================================
    final_stats = get_color_stats_from_image(
        reduced_img,
        color_book
    )

    print_color_table(final_stats)


    # =====================================================
    # 10. CREATE CARD GRID
    # =====================================================
    card_grid = create_card_grid(mapped_indices, palette_rgb)
    grid_img = card_grid.astype(np.uint8)


    # =====================================================
    # 11. PERFORMANCE EVALUATION
    # =====================================================
    mse_val, psnr_val, ssim_val, deltaE_val = evaluate(img_small, grid_img)

    print("\nGrid vs Resized Image Metrics")
    print(f"MSE   : {mse_val:.4f}")
    print(f"PSNR  : {psnr_val:.4f} dB")
    print(f"SSIM  : {ssim_val:.4f}")
    print(f"ΔE    : {deltaE_val:.4f}")

    report_filename = (
        os.path.splitext(os.path.basename(IMAGE_PATH))[0]
        + "_report.txt"
    )

    report_path = os.path.join(PERFORMANCE_DIR, report_filename)

    image_name = os.path.basename(IMAGE_PATH)

    save_report(
        report_path,
        image_name,
        img_small.shape,
        mse_val,
        psnr_val,
        ssim_val,
        deltaE_val
    )


    # =====================================================
    # 12. SPLIT GRID INTO PLATES
    # =====================================================
    plates = split_into_plates(
    mapped_indices,
    cards_per_row,
    cards_per_column
    )


    # =====================================================
    # 13. DISPLAY GRID
    # =====================================================
    display_card_grid(card_grid)


    # =====================================================
    # 14. EXPORT PLATES TO EXCEL
    # =====================================================
    plate_dir = os.path.join(BASE_DIR, "plate_sheet")
    os.makedirs(plate_dir, exist_ok=True)

    plate_filename = (
        os.path.splitext(os.path.basename(IMAGE_PATH))[0]
        + "_plates.xlsx"
    )

    plate_path = os.path.join(plate_dir, plate_filename)

    export_plates_to_excel(
        plates,
        cards_per_row,
        cards_per_column,
        plate_path
    )


    # =====================================================
    # 15. SAVE GRID IMAGE
    # =====================================================
    os.makedirs(OUTPUT_DIR, exist_ok=True)

    grid_filename = (
        os.path.splitext(os.path.basename(IMAGE_PATH))[0]
        + "_grid.png"
    )

    grid_path = os.path.join(OUTPUT_DIR, grid_filename)

    save_image(card_grid, grid_path)


    # =====================================================
    # 16. INTERACTIVE PLATE VIEWER
    # =====================================================
    while True:

        cmd = input(
            "\nEnter plate (A1 etc), 'full', 'nogrid', or 'quit': "
        ).strip()

        if cmd.lower() == "quit":
            break

        elif cmd.lower() == "nogrid":
            display_card_grid(card_grid, False)

        elif cmd.lower() == "full":
            display_card_grid(card_grid, True)

        else:
            display_specific_plate(plates, color_book, cmd)